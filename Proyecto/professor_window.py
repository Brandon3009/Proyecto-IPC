import tkinter as tk
import openpyxl
from tkinter import messagebox
from tkinter import simpledialog
from tkinter import ttk
from course_manager import CourseManager

class ProfessorWindow:
    def __init__(self, root, username):
        self.root = root
        self.username = username
        self.course_manager = CourseManager()
        self.courses = self.course_manager.get_courses_by_professor(username)
        self.selected_course = None  # Almacenar el curso seleccionado

    def show(self):
        self.window = tk.Toplevel(self.root)
        self.window.title(f"Ventana de Profesor - {self.username}")

        welcome_label = tk.Label(self.window, text=f"Bienvenido, Profesor {self.username}")
        welcome_label.pack(padx=10, pady=10)

        courses_label = tk.Label(self.window, text="Cursos a cargo:")
        courses_label.pack(padx=10, pady=5)

        self.course_listbox = tk.Listbox(self.window)
        self.course_listbox.pack(padx=10, pady=5)
        for course in self.courses:
            self.course_listbox.insert(tk.END, course["nombre"])

        edit_button = tk.Button(self.window, text="Editar Curso", command=self.editar_curso)
        edit_button.pack(padx=10, pady=5)

        view_grades_button = tk.Button(self.window, text="Ver Registro de Notas", command=self.ver_registro_notas)
        view_grades_button.pack(padx=10, pady=5)

        download_grades_button = tk.Button(self.window, text="Descargar Registro de Notas", command=self.descargar_registro_notas)
        download_grades_button.pack(padx=10, pady=5)

    def editar_curso(self):
        selected_course_index = self.course_listbox.curselection()
        if not selected_course_index:
            messagebox.showerror("Error", "Por favor, seleccione un curso para editar.")
            return

        self.selected_course = self.courses[selected_course_index[0] if selected_course_index else None]

        edit_course_window = tk.Toplevel(self.window)
        edit_course_window.title(f"Editar Curso - {self.selected_course['nombre']}")

        add_student_button = tk.Button(edit_course_window, text="Agregar Estudiante", command=self.agregar_estudiante)
        add_student_button.pack(padx=10, pady=5)

        remove_student_button = tk.Button(edit_course_window, text="Eliminar Estudiante", command=self.eliminar_estudiante)
        remove_student_button.pack(padx=10, pady=5)

        grade_student_button = tk.Button(edit_course_window, text="Calificar Estudiante", command=self.calificar_estudiante)
        grade_student_button.pack(padx=10, pady=5)

    def agregar_estudiante(self):
        student_name = simpledialog.askstring("Agregar Estudiante", "Ingrese el nombre del estudiante:")
        if student_name:
            if self.selected_course:
                self.selected_course['students'].append({'nombre': student_name, 'calificacion': None})
                messagebox.showinfo("Estudiante Agregado", f"Se agregó al estudiante: {student_name}")

    def eliminar_estudiante(self):
        if not self.selected_course:
            return

        selected_student_index = self.selected_course['students_listbox'].curselection()
        if not selected_student_index:
            messagebox.showerror("Error", "Por favor, seleccione un estudiante para eliminar.")
            return

        selected_student_index = selected_student_index[0]
        selected_student = self.selected_course['students'][selected_student_index]

        self.selected_course['students'].pop(selected_student_index)
        self.selected_course['students_listbox'].delete(selected_student_index)
        messagebox.showinfo("Estudiante Eliminado", f"Se eliminó al estudiante: {selected_student['nombre']}")

    def calificar_estudiante(self):
        if not self.selected_course:
            return

        selected_student_index = self.selected_course['students_listbox'].curselection()
        if not selected_student_index:
            messagebox.showerror("Error", "Por favor, seleccione un estudiante para calificar.")
            return

        selected_student_index = selected_student_index[0]
        selected_student = self.selected_course['students'][selected_student_index]

        grade = simpledialog.askfloat("Calificar Estudiante", f"Ingrese la calificación para {selected_student['nombre']}:")

        if grade is not None:
            selected_student['calificacion'] = grade
            self.selected_course['students_listbox'].delete(selected_student_index)
            self.selected_course['students_listbox'].insert(selected_student_index, f"{selected_student['nombre']} - {grade}")
            messagebox.showinfo("Calificación Registrada", f"Calificación de {selected_student['nombre']} actualizada: {grade}")

    def ver_registro_notas(self):
        if not self.selected_course:
            return

        view_grades_window = tk.Toplevel(self.window)
        view_grades_window.title(f"Registro de Notas - {self.selected_course['nombre']}")

        # Implementa la lógica para mostrar y editar el registro de notas aquí

    def descargar_registro_notas(self):
        if not self.selected_course:
            return

        workbook = openpyxl.Workbook()
        sheet = workbook.active

        row_index = 1
        sheet.cell(row=row_index, column=1, value="Nombre del Estudiante")
        sheet.cell(row=row_index, column=2, value="Calificación")

        for student in self.selected_course['students']:
            row_index += 1
            sheet.cell(row=row_index, column=1, value=student['nombre'])
            sheet.cell(row=row_index, column=2, value=student['calificacion'])

        workbook.save(f"{self.selected_course['nombre']}_registro_notas.xlsx")

if __name__ == "__main__":
    root = tk.Tk()
    professor_window = ProfessorWindow(root, "NombreProfesor")
    professor_window.show()
