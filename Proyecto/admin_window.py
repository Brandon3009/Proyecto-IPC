import tkinter as tk
from tkinter import ttk, messagebox, Listbox, Entry  # Agregar Listbox y Entry
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
import hashlib
import re  # Agregar formato 24h

class AdminWindow:

    def cerrar_admin(self):
        self.root.destroy()  # Cierra la ventana de administrador
        # Abre la ventana de inicio de sesión
        from login_window import LoginWindow
        login_window = LoginWindow()

    def __init__(self, root):
        self.root = root
        self.ventana_admin_cursos = None  # Agregar este atributo
        self.lista_usuarios = []  # Lista de usuarios registrados
        self.lista_cursos = []    # Lista de cursos
        self.lista_registros = []  # Estructura de datos para almacenar información de usuarios, cursos y notas

        # Llama para cargar los usuarios y registros al iniciar
        self.cargar_usuarios_desde_archivo()
        self.cargar_registros_desde_excel()
        self.cargar_cursos_desde_excel()  # Agrega esta línea para cargar los cursos

        # Agrega un botón de "Cerrar Sesión" en la esquina superior derecha
        boton_cerrar_sesion = ttk.Button(self.root, text="Cerrar Sesión", command=self.cerrar_admin)
        boton_cerrar_sesion.place(x=500, y=10)
        
        # Botones para administrar usuarios y cursos
        self.boton_administrar_usuarios = tk.Button(root, text="Administrar Usuarios", command=self.administrar_usuarios)
        self.boton_administrar_usuarios.pack(padx=10, pady=10)

        self.boton_administrar_cursos = tk.Button(root, text="Administrar Cursos", command=self.administrar_cursos)
        self.boton_administrar_cursos.pack(padx=10, pady=10)

        # Botón para ver el listado de notas
        self.boton_ver_notas = tk.Button(root, text="Ver Listado de Notas", command=self.ver_listado_notas)
        self.boton_ver_notas.pack(padx=10, pady=10)

        # Botón para registrar un nuevo profesor
        boton_registrar_profesor = tk.Button(root, text="Registrar Nuevo Profesor", command=self.registrar_profesor)
        boton_registrar_profesor.pack(padx=10, pady=10)

    def registrar_profesor(self):
        ventana_registro_profesor = tk.Toplevel(self.root)
        ventana_registro_profesor.title("Registrar Profesor")
        ventana_registro_profesor.geometry("400x400")
        ventana_registro_profesor.resizable(0, 0)
        ventana_registro_profesor.config(bd=10,bg="lavender blush")

        label_nombre_profesor = tk.Label(ventana_registro_profesor, text="Nombre del Profesor:", bg="lavender blush")
        label_nombre_profesor.pack(padx=10, pady=5)
        entry_nombre_profesor = tk.Entry(ventana_registro_profesor)
        entry_nombre_profesor.pack(padx=10, pady=5)

        label_apellido_profesor = tk.Label(ventana_registro_profesor, text="Apellido del Profesor:", bg="lavender blush")
        label_apellido_profesor.pack(padx=10, pady=5)
        entry_apellido_profesor = tk.Entry(ventana_registro_profesor)
        entry_apellido_profesor.pack(padx=10, pady=5)

        label_cui_profesor = tk.Label(ventana_registro_profesor, text="CUI del Profesor:", bg="lavender blush")
        label_cui_profesor.pack(padx=10, pady=5)
        entry_cui_profesor = tk.Entry(ventana_registro_profesor)
        entry_cui_profesor.pack(padx=10, pady=5)

        label_contrasena = tk.Label(ventana_registro_profesor, text="Contraseña:", bg="lavender blush")
        label_contrasena.pack(padx=10, pady=5)
        entry_contrasena = tk.Entry(ventana_registro_profesor, show="*")  # Para ocultar la contraseña
        entry_contrasena.pack(padx=10, pady=5)

        label_confirmar_contrasena = tk.Label(ventana_registro_profesor, text="Confirmar Contraseña:", bg="lavender blush")
        label_confirmar_contrasena.pack(padx=10, pady=5)
        entry_confirmar_contrasena = tk.Entry(ventana_registro_profesor, show="*")  # Para ocultar la contraseña
        entry_confirmar_contrasena.pack(padx=10, pady=5)

        var_tipo_usuario = tk.StringVar()
        var_tipo_usuario.set("Profesor")

        def guardar_profesor():
            nombre = entry_nombre_profesor.get()
            apellido = entry_apellido_profesor.get()
            cui = entry_cui_profesor.get()
            contrasena = entry_contrasena.get()
            confirmar_contrasena = entry_confirmar_contrasena.get()
            tipo_usuario = var_tipo_usuario.get()
    
            if nombre and apellido and cui and contrasena and contrasena == confirmar_contrasena:
                      # Verificar si el CUI ya existe
                if self.verificar_cui_existente(cui):
                    messagebox.showerror("Error", "El CUI ya existe. Ingrese un CUI único.")
                else:
                        # Hash de la contraseña
                    hashed_password = hashlib.sha256(contrasena.encode()).hexdigest()

                       # Guardar los datos del profesor en el archivo profesores.txt
                    with open("profesores.txt", "a") as archivo:
                        archivo.write(f"{nombre}:{apellido}:{cui}:{hashed_password}:{tipo_usuario}\n")

                    messagebox.showinfo("Profesor Registrado", f"El profesor '{nombre}' ha sido registrado con éxito.")
                    ventana_registro_profesor.destroy()
            else:
                messagebox.showerror("Error", "Por favor, complete todos los campos correctamente.")

        boton_guardar_profesor = tk.Button(ventana_registro_profesor, text="Guardar Profesor", command=guardar_profesor)
        boton_guardar_profesor.pack(padx=10, pady=10)

    def verificar_cui_existente(self, cui):
        for usuario, bloqueado in self.lista_usuarios:
            if cui == usuario:
                return True
        return False

    def cargar_usuarios_desde_archivo(self):  # Cambiado el nombre de la función
        self.lista_usuarios = []  # Limpiar la lista antes de cargar usuarios

        try:
            with open("usuarios.txt", "r") as archivo:
                for linea in archivo:
                    partes = linea.strip().split(':')
                    if len(partes) >= 4:
                        usuario = partes[0]
                        bloqueado = "bloqueado" in partes[-1]  # Verificar si está bloqueado
                        self.lista_usuarios.append((usuario, bloqueado))
        except FileNotFoundError:
        # El archivo de usuarios no existe, la lista estará vacía
            pass

    def cargar_registros_desde_excel(self):
        try:
            workbook = openpyxl.load_workbook("registros.xlsx")
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                nombre, apellido, curso, nota = row
                registro = {
                    "nombre": nombre,
                    "apellido": apellido,
                    "curso": curso,
                    "nota": nota
                }
                self.lista_registros.append(registro)
        except FileNotFoundError:
        # El archivo de registros no existe, la lista estará vacía
            pass

    def ver_listado_notas(self):
        notas_workbook = Workbook()
        notas_sheet = notas_workbook.active
        notas_sheet.title = "Listado de Notas"
        notas_sheet.append(["Nombre", "Apellido", "Curso", "Nota"])

        for registro in self.lista_registros:
            nombre = registro['nombre']
            apellido = registro['apellido']
            curso = registro['curso']
            nota = registro['nota']
            notas_sheet.append([nombre, apellido, curso, nota])

        notas_workbook.save("listado_notas.xlsx")

        messagebox.showinfo("Listado de Notas", "El listado de notas se ha guardado en 'listado_notas.xlsx'.")

    def administrar_usuarios(self):
        # Llama a la función para cargar los usuarios antes de abrir la ventana
        self.cargar_usuarios_desde_archivo()

        ventana_admin_usuarios = tk.Toplevel(self.root)
        ventana_admin_usuarios.title("Administrar Usuarios")
        ventana_admin_usuarios.geometry("400x350")  # Ajustar el tamaño de la ventana
        ventana_admin_usuarios.resizable(0, 0)
        ventana_admin_usuarios.config(bd=10,bg="lightcyan")

        var_usuario_seleccionado = tk.StringVar(ventana_admin_usuarios)

        # Crear un Entry para la búsqueda de usuarios
        entry_busqueda = Entry(ventana_admin_usuarios,bg="lightcyan")
        entry_busqueda.pack(padx=10, pady=5)
        entry_busqueda.insert(0, "Buscar usuario")

        lista_usuarios = Listbox(ventana_admin_usuarios, bg="lightcyan")
        lista_usuarios.pack(padx=10, pady=5)

        if self.lista_usuarios:
            for usuario, bloqueado in self.lista_usuarios:
                bloqueo = "Bloqueado" if bloqueado else ""
                lista_usuarios.insert(tk.END, f"{usuario} - {bloqueo}")

        def actualizar_lista_usuarios():
            lista_usuarios.delete(0, tk.END)
            for usuario, bloqueado in self.lista_usuarios:
                bloqueo = "Bloqueado" if bloqueado else ""
                lista_usuarios.insert(tk.END, f"{usuario} - {bloqueo}")

        # Agregar una función para buscar usuarios
        def buscar_usuario():
            texto_busqueda = entry_busqueda.get().strip().lower()
            usuarios_coincidentes = []
            for usuario, bloqueado in self.lista_usuarios:
                if texto_busqueda in usuario.lower():
                    usuarios_coincidentes.append((usuario, bloqueado))
            self.lista_usuarios = usuarios_coincidentes
            actualizar_lista_usuarios()

        entry_busqueda.bind("<Return>", lambda event=None: buscar_usuario())

        def bloquear_usuario_seleccionado():
            usuario_seleccionado = lista_usuarios.get(lista_usuarios.curselection())
            usuario_seleccionado = usuario_seleccionado.split(" - ")[0]
            bloqueado = False  # Variable para determinar si el usuario ya está bloqueado

            with open("usuarios.txt", "r") as archivo:
                lineas = archivo.readlines()

            with open("usuarios.txt", "w") as archivo:
                for linea in lineas:
                    partes = linea.strip().split(':')
                    if len(partes) >= 4 and partes[0] == usuario_seleccionado:
                        if "bloqueado" in partes[-1]:
                            bloqueado = True
                        else:
                            partes[-1] += ",bloqueado"
                        linea = ":".join(partes) + '\n'
                    archivo.write(linea)

                # Actualizar la lista de usuarios
            self.cargar_usuarios_desde_archivo()
            actualizar_lista_usuarios()

            if bloqueado:
                messagebox.showinfo("Usuario Bloqueado", f"El usuario '{usuario_seleccionado}' ya estaba bloqueado.")
            else:
                messagebox.showinfo("Usuario Bloqueado", f"El usuario '{usuario_seleccionado}' ha sido bloqueado.")

        def desbloquear_usuario_seleccionado():
            usuario_seleccionado = lista_usuarios.get(lista_usuarios.curselection())
            usuario_seleccionado = usuario_seleccionado.split(" - ")[0]
            desbloqueado = False  # Variable para determinar si el usuario ya estaba desbloqueado

            with open("usuarios.txt", "r") as archivo:
                lineas = archivo.readlines()

            with open("usuarios.txt", "w") as archivo:
                for linea in lineas:
                    partes = linea.strip().split(':')
                    if len(partes) >= 4 and partes[0] == usuario_seleccionado:
                        if ",bloqueado" in partes[-1]:
                            partes[-1] = partes[-1].replace(",bloqueado", "")
                            desbloqueado = True
                        linea = ":".join(partes) + '\n'
                    archivo.write(linea)

                # Actualizar la lista de usuarios
            self.cargar_usuarios_desde_archivo()
            actualizar_lista_usuarios()

            if desbloqueado:
                messagebox.showinfo("Usuario Desbloqueado", f"El usuario '{usuario_seleccionado}' ha sido desbloqueado.")
            else:
                messagebox.showinfo("Usuario Desbloqueado", f"El usuario '{usuario_seleccionado}' ya estaba desbloqueado.")

        boton_bloquear_usuario = tk.Button(ventana_admin_usuarios, text="Bloquear Usuario", command=bloquear_usuario_seleccionado)
        boton_bloquear_usuario.pack(padx=10, pady=10)

        boton_desbloquear_usuario = tk.Button(ventana_admin_usuarios, text="Desbloquear Usuario", command=desbloquear_usuario_seleccionado)
        boton_desbloquear_usuario.pack(padx=10, pady=10)

    def cargar_cursos_desde_excel(self):
        try:
            workbook = openpyxl.load_workbook("cursos.xlsx")
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                nombre_curso, descripcion_curso, costo_curso, horario_curso, codigo_curso, cupo_curso, catedratico_curso = row
                curso = {
                    "nombre": nombre_curso,
                    "descripcion": descripcion_curso,
                    "costo": costo_curso,
                    "horario": horario_curso,
                    "codigo": codigo_curso,
                    "cupo": cupo_curso,
                    "catedratico": catedratico_curso
                }
                self.lista_cursos.append(curso)
        except FileNotFoundError:
            # El archivo de cursos no existe, la lista estará vacía
            pass


    def administrar_cursos(self):
        # Cargar los cursos antes de abrir la ventana
        self.cargar_cursos_desde_excel()

        self.ventana_admin_cursos = tk.Toplevel(self.root)
        self.ventana_admin_cursos.title("Administrar Cursos")
        self.ventana_admin_cursos.geometry("300x300")
        self.ventana_admin_cursos.resizable(0, 0)
        self.ventana_admin_cursos.config(bd=10, bg="pale turquoise")

        boton_crear_curso = tk.Button(self.ventana_admin_cursos, text="Crear Curso", command=self.crear_curso, bg="pale turquoise")
        boton_crear_curso.pack(padx=10, pady=10)

        boton_ver_cursos = tk.Button(self.ventana_admin_cursos, text="Ver Listado de Cursos", command=self.ver_listado_cursos, bg="pale turquoise")
        boton_ver_cursos.pack(padx=10, pady=10)

        boton_borrar_curso = tk.Button(self.ventana_admin_cursos, text="Borrar Curso", command=self.borrar_curso, bg="pale turquoise")
        boton_borrar_curso.pack(padx=10, pady=10)

        boton_borrar_todos_los_cursos = tk.Button(self.ventana_admin_cursos, text="Borrar Todos los Cursos", command=self.borrar_todos_los_cursos, bg="pale turquoise")
        boton_borrar_todos_los_cursos.pack(padx=10, pady=10)

    def crear_curso(self):
        ventana_creacion_curso = tk.Toplevel(self.root)
        ventana_creacion_curso.title("Crear Curso")
        ventana_creacion_curso.geometry("400x500")  # Ajusta el tamaño de la ventana
        ventana_creacion_curso.resizable(0, 0)
        ventana_creacion_curso.config(bd=10, bg="pale turquoise")

        label_nombre_curso = tk.Label(ventana_creacion_curso, text="Nombre del Curso:", bg="pale turquoise")
        label_nombre_curso.pack(padx=10, pady=5)
        entry_nombre_curso = tk.Entry(ventana_creacion_curso)
        entry_nombre_curso.pack(padx=10, pady=5)

        label_descripcion_curso = tk.Label(ventana_creacion_curso, text="Descripción del Curso:", bg="pale turquoise")
        label_descripcion_curso.pack(padx=10, pady=5)
        entry_descripcion_curso = tk.Entry(ventana_creacion_curso)
        entry_descripcion_curso.pack(padx=10, pady=5)

      # Agregar campos para costo, horario, código, cupo y catedrático
        label_costo = tk.Label(ventana_creacion_curso, text="Costo del Curso:", bg="pale turquoise")
        label_costo.pack(padx=10, pady=5)
        entry_costo = tk.Entry(ventana_creacion_curso)
        entry_costo.pack(padx=10, pady=5)

        # Lista de horarios disponibles
        horarios_disponibles = [
            "07:00 - 09:30",
            "09:35 - 10:30",
            "10:35 - 11:30",
            "11:35 - 13:00",
            "13:05 - 15:00",
            "15:05 - 16:00",
            "16:05 - 17:00",
            "17:05 - 18:00",
            "18:05 - 20:00",
            "20:05 - 20:50",
            "21:00 - 22:00",
        ]

         # Etiqueta para seleccionar el horario
        label_horario = tk.Label(ventana_creacion_curso, text="Horario del Curso:",bg="pale turquoise")
        label_horario.pack(padx=10, pady=5)

         # Combobox para seleccionar el horario
        var_horario = tk.StringVar(ventana_creacion_curso)
        var_horario.set(horarios_disponibles[0])  # Establece el primer horario como valor predeterminado

        combo_horario = ttk.Combobox(ventana_creacion_curso, textvariable=var_horario, values=horarios_disponibles)
        combo_horario.pack(padx=10, pady=5)

        # Función para validar el formato de horario
        def validar_horario(horario):
              # Utiliza una expresión regular para validar el formato HH:MM - HH:MM
            return re.match(r'^([01]\d|2[0-3]):[0-5]\d - ([01]\d|2[0-3]):[0-5]\d$', horario) is not None

        label_codigo = tk.Label(ventana_creacion_curso, text="Código del Curso:", bg="pale turquoise")
        label_codigo.pack(padx=10, pady=5)
        entry_codigo = tk.Entry(ventana_creacion_curso)
        entry_codigo.pack(padx=10, pady=5)

        label_cupo = tk.Label(ventana_creacion_curso, text="Cupo del Curso:", bg="pale turquoise")
        label_cupo.pack(padx=10, pady=5)
        entry_cupo = tk.Entry(ventana_creacion_curso)
        entry_cupo.pack(padx=10, pady=5)

        # Agregar un campo para seleccionar el catedrático
        label_catedratico = tk.Label(ventana_creacion_curso, text="Catedrático del Curso:", bg="pale turquoise")
        label_catedratico.pack(padx=10, pady=5)

        # Crear una lista para almacenar los nombres de los profesores
        nombres_profesores = []

            # Leer los nombres de los profesores desde "profesores.txt" y evitar repeticiones
        with open("profesores.txt", "r") as archivo:
            for linea in archivo:
                partes = linea.strip().split(':')
                if len(partes) >= 2:
                    nombre_profesor = partes[1]
                    if nombre_profesor not in nombres_profesores:
                        nombres_profesores.append(nombre_profesor)

           # Crear una variable para seleccionar al catedrático
        var_catedratico = tk.StringVar(ventana_creacion_curso)
        var_catedratico.set(nombres_profesores[0] if nombres_profesores else "")  # Establece el primer profesor como valor predeterminado

          # Caja de selección para elegir el catedrático
        combo_catedratico = ttk.Combobox(ventana_creacion_curso, textvariable=var_catedratico, values=nombres_profesores)
        combo_catedratico.pack(padx=10, pady=5)
        
        def guardar_curso():
            nombre_curso = entry_nombre_curso.get()
            descripcion_curso = entry_descripcion_curso.get()
            costo_curso = entry_costo.get()
            horario_curso = var_horario.get()
            codigo_curso = entry_codigo.get()
            cupo_curso = entry_cupo.get()
            catedratico_curso = var_catedratico.get()

              # Validar el formato del horario
            if not validar_horario(horario_curso):
                messagebox.showerror("Error", "Formato de horario no válido. Debe ser 'HH:MM - HH:MM'")
                return

            if nombre_curso and descripcion_curso and costo_curso and horario_curso and codigo_curso and cupo_curso and catedratico_curso:
                # Intenta cargar el archivo existente o crear uno nuevo si no existe
                try:
                    workbook = openpyxl.load_workbook("cursos.xlsx")
                    sheet = workbook.active
                except FileNotFoundError:
                    workbook = Workbook()
                    sheet = workbook.active
                    sheet.append(["Nombre del Curso", "Descripción del Curso", "Costo", "Horario", "Código", "Cupo", "Catedrático"])

                    # Agrega la información del nuevo curso sin repetir el encabezado
                sheet.append([nombre_curso, descripcion_curso, costo_curso, horario_curso, codigo_curso, cupo_curso, catedratico_curso])

                 # Guarda el archivo
                workbook.save("cursos.xlsx")

                with open("cursos.txt", "a") as archivo_texto:
                    archivo_texto.write(f"{nombre_curso},{descripcion_curso},{costo_curso},{horario_curso},{codigo_curso},{cupo_curso},{catedratico_curso}\n")

                messagebox.showinfo("Curso Creado", f"El curso '{nombre_curso}' ha sido creado con éxito.")
                ventana_creacion_curso.destroy()
            else:
                messagebox.showerror("Error", "Por favor, complete todos los campos.")

        boton_guardar_curso = tk.Button(ventana_creacion_curso, text="Guardar Curso", command=guardar_curso)
        boton_guardar_curso.pack(padx=10, pady=10)

    def borrar_curso(self):
        ventana_borrar_curso = tk.Toplevel(self.ventana_admin_cursos)
        ventana_borrar_curso.title("Borrar Curso")
        ventana_borrar_curso.geometry("300x150")
        ventana_borrar_curso.resizable(0, 0)
        ventana_borrar_curso.config(bd=10)

          # Crear una lista de nombres de cursos
        nombres_cursos = [curso["nombre"] for curso in self.lista_cursos]

        if not nombres_cursos:
            messagebox.showinfo("Info", "No hay cursos para borrar.")
            ventana_borrar_curso.destroy()
            return

        var_curso_a_borrar = tk.StringVar(ventana_borrar_curso)
        var_curso_a_borrar.set(nombres_cursos[0])  # Establece el primer curso como valor predeterminado

          # Crear un menú desplegable para seleccionar el curso a borrar
        label_curso_a_borrar = tk.Label(ventana_borrar_curso, text="Selecciona un curso a borrar:")
        label_curso_a_borrar.pack(padx=10, pady=5)

        combo_cursos = ttk.Combobox(ventana_borrar_curso, textvariable=var_curso_a_borrar, values=nombres_cursos)
        combo_cursos.pack(padx=10, pady=5)

        def confirmar_borrar_curso():
            curso_seleccionado = var_curso_a_borrar.get()
            nombres_cursos.remove(curso_seleccionado)  # Elimina el curso de la lista de nombres
            combo_cursos["values"] = nombres_cursos  # Actualiza la lista de cursos disponibles

            if not curso_seleccionado:
                messagebox.showinfo("Info", "No hay cursos para borrar.")
                ventana_borrar_curso.destroy()
                return

                # Buscar el curso por nombre y eliminarlo de la lista de cursos
            curso_eliminado = None
            for curso in self.lista_cursos:
                if curso["nombre"] == curso_seleccionado:
                    curso_eliminado = curso
                    break

            if curso_eliminado:
                self.lista_cursos.remove(curso_eliminado)

                 # Actualizar el archivo de cursos
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.append(["Nombre del Curso", "Descripción del Curso", "Costo", "Horario", "Código", "Cupo", "Catedrático"])

                for curso in self.lista_cursos:
                    sheet.append([curso["nombre"], curso["descripcion"], curso["costo"], curso["horario"], curso["codigo"], curso["cupo"], curso["catedratico"]])

                workbook.save("cursos.xlsx")

                # Actualizar la lista de cursos disponibles para eliminar el curso seleccionado
                nombres_cursos.remove(curso_seleccionado)
                combo_cursos["values"] = nombres_cursos

                messagebox.showinfo("Borrar Curso", f"El curso '{curso_seleccionado}' ha sido borrado con éxito.")
            else:
                messagebox.showinfo("Info", "No hay cursos para borrar.")

            ventana_borrar_curso.destroy()

        boton_confirmar_borrar_curso = tk.Button(ventana_borrar_curso, text="Borrar Curso", command=confirmar_borrar_curso)
        boton_confirmar_borrar_curso.pack(padx=10, pady=5)

          # Agregar un botón para borrar todos los cursos

    def borrar_todos_los_cursos(self):
        if self.ventana_admin_cursos:
            if not self.lista_cursos:
                messagebox.showinfo("Info", "La lista de cursos está vacía.")
            else:
                respuesta = messagebox.askyesno("Borrar Todos los Cursos", "¿Estás seguro de que deseas borrar todos los cursos?")
                if respuesta:
                        # Limpiar la lista de cursos
                    self.lista_cursos.clear()

                       # Actualizar el archivo de cursos
                    workbook = openpyxl.Workbook()
                    sheet = workbook.active
                    sheet.append(["Nombre del Curso", "Descripción del Curso", "Costo", "Horario", "Código", "Cupo", "Catedrático"])
                    workbook.save("cursos.xlsx")

                        # Actualizar la lista de cursos disponibles
                    combo_cursos = ttk.Combobox(self.ventana_admin_cursos)
                    combo_cursos.pack(padx=10, pady=5)
                    combo_cursos['values'] = []

                    messagebox.showinfo("Borrar Cursos", "Todos los cursos han sido borrados con éxito.")
        else:
            messagebox.showinfo("Info", "La ventana 'Administrar Cursos' no ha sido inicializada.")

    def ver_listado_cursos(self):
       # Cargar los cursos antes de abrir la ventana
        self.cargar_cursos_desde_excel()

        if self.lista_cursos:
            cursos_workbook = Workbook()
            cursos_sheet = cursos_workbook.active
            cursos_sheet.title = "Listado de Cursos"
            cursos_sheet.append(["Nombre del Curso", "Descripción del Curso", "Costo", "Horario", "Código", "Cupo", "Catedrático"])

            for curso in self.lista_cursos:
                nombre_curso = curso.get("nombre", "")
                descripcion_curso = curso.get("descripcion", "")
                costo = curso.get("costo", "")
                horario = curso.get("horario", "")
                codigo = curso.get("codigo", "")
                cupo = curso.get("cupo", "")
                catedratico = curso.get("catedratico", "")
                cursos_sheet.append([nombre_curso, descripcion_curso, costo, horario, codigo, cupo, catedratico])

            cursos_workbook.save("listado_cursos.xlsx")

            messagebox.showinfo("Listado de Cursos", "El listado de cursos se ha guardado en 'listado_cursos.xlsx'.")
        else:
            messagebox.showerror("Error", "No hay cursos disponibles para mostrar.")

    def ver_listado_notas(self):
        # Crear un archivo Excel para listar notas
        notas_workbook = Workbook()
        notas_sheet = notas_workbook.active
        notas_sheet.title = "Listado de Notas"
        notas_sheet.append(["Nombre", "Apellido", "Curso", "Nota"])

        # Iterar sobre los registros y agregar los datos al archivo Excel
        for registro in self.lista_registros:
            nombre = registro['nombre']
            apellido = registro['apellido']
            curso = registro['curso']
            nota = registro['nota']
            notas_sheet.append([nombre, apellido, curso, nota])

        # Guardar el archivo Excel de notas
        notas_workbook.save("listado_notas.xlsx")

        messagebox.showinfo("Listado de Notas", "El listado de notas se ha guardado en 'listado_notas.xlsx'.")

if __name__ == "__main__":
    ventana_principal = tk.Tk()
    ventana_principal.title("Ventana Administrador")
    ventana_principal.geometry("600x400")
    ventana_principal.config(bd=10, bg="wheat")

    admin_window = AdminWindow(ventana_principal)

    ventana_principal.mainloop()
    